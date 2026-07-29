from pathlib import Path
import json
import re
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "data" / "domains"
OUT_DIR.mkdir(parents=True, exist_ok=True)

FILES = [
    {
        "domain_id": "fpa",
        "domain_label": "FP&A",
        "domain_title": "Planificación y análisis financiero / FP&A",
        "source": ROOT / "F3M_FP&A_assessment_ready_AI_v1.xlsx",
        "output": OUT_DIR / "fpa.json",
    },
    {
        "domain_id": "controlling",
        "domain_label": "Controlling",
        "domain_title": "Controlling",
        "source": ROOT / "F3M_Controlling_assessment_ready_AI_v0.xlsx",
        "output": OUT_DIR / "controlling.json",
    },
    {
        "domain_id": "transacciones",
        "domain_label": "Transacciones",
        "domain_title": "Transacciones",
        "source": ROOT / "F3M_Transaccional_assessment_ready_AI.xlsx",
        "output": OUT_DIR / "transacciones.json",
    },
    {
        "domain_id": "finanzas-negocio",
        "domain_label": "Finanzas de negocio",
        "domain_title": "Finanzas de negocio",
        "source": ROOT / "F3M_Finanzas_Negocio_BU_assessment_ready_AI_v0.xlsx",
        "output": OUT_DIR / "finanzas-negocio.json",
    },
    {
        "domain_id": "auditoria-interna",
        "domain_label": "Auditoría Interna",
        "domain_title": "Auditoría Interna",
        "source": ROOT / "F3M_Auditoria_Interna_assessment_ready_AI_v0.xlsx",
        "output": OUT_DIR / "auditoria-interna.json",
    },
    {
        "domain_id": "finanzas-estrategicas",
        "domain_label": "Finanzas Estratégicas",
        "domain_title": "Finanzas Estratégicas",
        "source": ROOT / "F3M_Finanzas_Estrategicas_assessment_ready_AI_v0.xlsx",
        "output": OUT_DIR / "finanzas-estrategicas.json",
    },
    {
        "domain_id": "relacion-inversores",
        "domain_label": "Relación con Inversores",
        "domain_title": "Relación con Inversores",
        "source": ROOT / "F3M_Relacion_Inversores_assessment_ready_AI_v0.xlsx",
        "output": OUT_DIR / "relacion-inversores.json",
    },
    {
        "domain_id": "tesoreria",
        "domain_label": "Tesorería",
        "domain_title": "Tesorería",
        "source": ROOT / "F3M_Tesoreria_assessment_ready_AI_v0.xlsx",
        "output": OUT_DIR / "tesoreria.json",
    },
    {
        "domain_id": "fiscal",
        "domain_label": "Fiscal",
        "domain_title": "Fiscal",
        "source": ROOT / "F3M_Fiscal_assessment_ready_AI_v0.xlsx",
        "output": OUT_DIR / "fiscal.json",
    },
]


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def to_number(value):
    if value is None or value == "":
        return None
    try:
        number = float(value)
        if number == 0:
            return None
        if number.is_integer():
            return int(number)
        return number
    except Exception:
        return None


def split_lines(value):
    text = clean(value)
    if not text:
        return []
    return [line.strip() for line in re.split(r"\n+", text) if line.strip()]


def sheet_rows(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [clean(cell) for cell in rows[0]]
    output = []

    for row in rows[1:]:
        if not any(cell not in (None, "") for cell in row):
            continue

        record = {}
        for index, header in enumerate(headers):
            if header:
                record[header] = row[index] if index < len(row) else None
        output.append(record)

    return output


def build_ai_map(workbook):
    if "AI Overlay" not in workbook.sheetnames:
        return {}

    rows = sheet_rows(workbook["AI Overlay"])
    ai_map = {}

    for row in rows:
        capacidad = clean(row.get("Capacidad"))
        subcapacidad = clean(row.get("Subcapacidad"))

        if not capacidad:
            continue

        key = f"{capacidad}||{subcapacidad}" if subcapacidad else capacidad

        ai_map[key] = {
            "capacidad": capacidad,
            "subcapacidad": subcapacidad,
            "cases": clean(row.get("Casos de uso AI asociados")),
            "advanced": clean(row.get("Aplicación en nivel avanzado")),
            "source": clean(row.get("Fuente")),
        }

    return ai_map


def find_ai_for_row(row, ai_map):
    capacidad = clean(row.get("Capacidad"))
    subcapacidad = clean(row.get("Subcapacidad"))

    direct_key = f"{capacidad}||{subcapacidad}"
    if direct_key in ai_map:
        return ai_map[direct_key]

    if capacidad in ai_map:
        return ai_map[capacidad]

    # FP&A AI Overlay uses ranges like 1.1-1.4 by capacidad.
    for value in ai_map.values():
        if value.get("capacidad") == capacidad:
            return value

    ai_cases = clean(row.get("Casos de uso AI asociados"))
    if ai_cases:
        return {
            "capacidad": capacidad,
            "subcapacidad": subcapacidad,
            "cases": ai_cases,
            "advanced": clean(row.get("Aplicación en nivel avanzado")),
            "source": "",
        }

    return {
        "capacidad": capacidad,
        "subcapacidad": subcapacidad,
        "cases": "",
        "advanced": "",
        "source": "",
    }


def make_id(domain_id, subcapacidad, index):
    raw = clean(subcapacidad).lower()
    prefix_match = re.match(r"^(\d+)\.(\d+)", raw)
    if prefix_match:
        return f"{domain_id}-{prefix_match.group(1)}-{prefix_match.group(2)}"

    slug = re.sub(r"[^a-z0-9]+", "-", raw).strip("-")
    return f"{domain_id}-{slug or index + 1}"


def convert_domain(config):
    wb = load_workbook(config["source"], data_only=True)

    assessment_rows = sheet_rows(wb["Assessment"])
    ai_map = build_ai_map(wb)

    items = []

    for index, row in enumerate(assessment_rows):
        capacidad = clean(row.get("Capacidad"))
        subcapacidad = clean(row.get("Subcapacidad"))

        if not capacidad or not subcapacidad:
            continue

        ai = find_ai_for_row(row, ai_map)

        item = {
            "id": make_id(config["domain_id"], subcapacidad, index),
            "capacidad": capacidad,
            "subcapacidad": subcapacidad,
            "objetivo": clean(row.get("Objetivo de evaluación")),
            "maturity": {
                "1": clean(row.get("Nivel 1 - Inicial")),
                "2": clean(row.get("Nivel 2 - Estructurado")),
                "3": clean(row.get("Nivel 3 - Estandarizado")),
                "4": clean(row.get("Nivel 4 - Optimizado")),
                "5": clean(row.get("Nivel 5 - Avanzado/Referente")),
            },
            "scores": {
                "procesos": to_number(row.get("Score Procesos")),
                "tecnologia": to_number(row.get("Score Tecnología")),
                "organizacion": to_number(row.get("Score Organización")),
            },
            "preguntas": split_lines(row.get("Preguntas clave")),
            "evidencias": clean(row.get("Evidencias")),
            "iniciativaSugerida": clean(row.get("Iniciativa sugerida")),
            "comentario": clean(row.get("Comentarios / hallazgos")),
            "owner": "",
            "status": "No iniciado",
            "ai": {
                "cases": ai.get("cases", ""),
                "advanced": ai.get("advanced", ""),
                "source": ai.get("source", ""),
                "subcapacidad": ai.get("subcapacidad", ""),
            },
        }

        items.append(item)

    payload = {
        "meta": {
            "domainId": config["domain_id"],
            "domainLabel": config["domain_label"],
            "domainTitle": config["domain_title"],
            "sourceFile": config["source"].name,
            "targetMaturity": 4,
        },
        "subcapacities": items,
    }

    config["output"].write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    print(f"OK: {config['output']} -> {len(items)} subcapacidades")


def main():
    for config in FILES:
        convert_domain(config)


if __name__ == "__main__":
    main()